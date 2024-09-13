from django.http import HttpResponse
from django.shortcuts import render,redirect
from siyyana_app.models import *
from siyyana_app.forms import *
from django.contrib import messages
from siyyana_app.forms import ServicesAddForm
from django.contrib.auth.models import User,auth
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from accounts.models import *
from .utils import top_booked_employees
from django.db.models import Count
import openpyxl
from io import BytesIO
import pandas as pd
# Create your views here.

@login_required(login_url="siyyana_app:login")
def index(request):
    employee_count = CustomUser.objects.filter(user_type = 'Employee').count()
    user_count = CustomUser.objects.filter(user_type = 'User').count()
    booking_count = Booking.objects.all().count()
    services_count = Category.objects.all().count()
    sub_services_count = SubCategory.objects.all().count()
    top_services_count = TopCategory.objects.all().count()
    # Get the top booked employees
    top_booked_employees = CustomUser.objects.filter(user_type = 'Employee')[:5]
    top_booked_customers = CustomUser.objects.filter(user_type = 'User')[:5]

    context = {

        'employee_count':employee_count,
        'user_count':user_count,
        'booking_count':booking_count,
        'services_count':services_count,
        'sub_services_count':sub_services_count,
        'top_services_count':top_services_count,
        'top_booked_employees': top_booked_employees, 
        'top_booked_customers': top_booked_customers

    }
    return render(request,'index.html',context)






def login(request):
    if request.method == 'POST':
        username_or_email = request.POST.get('username_or_email')
        password = request.POST.get('password')
        user = auth.authenticate(request, username=username_or_email, password=password)
        if user is not None:
            auth.login(request, user)
            request.session['username'] = username_or_email
            return redirect('siyyana_app:index')
        else:
            messages.error(request, 'Invalid credential..')
            return redirect('siyyana_app:login')
    return render(request, 'login.html')




def admin_logout(request):
    logout(request)
    return redirect("siyyana_app:login")


@login_required(login_url="siyyana_app:login")
def services(request):
    data = Category.objects.all()
    context = {'data':data}
    return render(request,'services.html',context)


@login_required(login_url="siyyana_app:login")
def add_services(request):
    if request.method == 'POST':
        form = ServicesAddForm(request.POST,request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request,'Service added successfully')
            return redirect('siyyana_app:services')
    else:
        form = ServicesAddForm()
    context = {'form':form}
    return render(request,'add-services.html',context)

@login_required(login_url="siyyana_app:login")
def edit_services(request,id):
    update = Category.objects.get(id=id)
    form = ServicesAddForm(instance=update)
    if request.method == 'POST':
        form = ServicesAddForm(request.POST,request.FILES,instance=update)
        if form.is_valid():
            form.save()
            messages.success(request,'Service added successfully')
            return redirect('siyyana_app:services')
    context = {'form':form}
    return render(request,'add-services.html',context)



@login_required(login_url="siyyana_app:login")
def delete_services(request,id):
    delete = Category.objects.get(id=id)
    delete.delete()
    return redirect('siyyana_app:services')




@login_required(login_url="siyyana_app:login")
def sub_services(request):
    data = SubCategory.objects.all().order_by('-id')
    context = {'data':data}
    return render(request,'subservices.html',context)


@login_required(login_url="siyyana_app:login")
def add_subservices(request):
    if request.method == 'POST':
        form = SubServicesAddForm(request.POST,request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request,'SubService added successfully')
            return redirect('siyyana_app:sub_services')
    else:
        form = SubServicesAddForm()
    context = {'form':form}
    return render(request,'add-sub-services.html',context)


@login_required(login_url="siyyana_app:login")
def edit_subservices(request,id):
    update = SubCategory.objects.get(id=id)
    form = SubServicesAddForm(instance=update)
    if request.method == 'POST':
        form = SubServicesAddForm(request.POST,request.FILES,instance=update)
        if form.is_valid():
            form.save()
            messages.success(request,'SubService added successfully')
            return redirect('siyyana_app:sub_services')
    context = {'form':form}
    return render(request,'add-sub-services.html',context)


@login_required(login_url="siyyana_app:login")
def delete_subservices(request,id):
    delete = SubCategory.objects.get(id=id)
    delete.delete()
    return redirect('siyyana_app:sub_services')



@login_required(login_url="siyyana_app:login")
def top_services(request):
    data = TopCategory.objects.all().order_by('-id')
    context = {'data':data}
    return render(request,'top-services.html',context)


@login_required(login_url="siyyana_app:login")
def add_topservices(request):
    if request.method == 'POST':
        form = TopServicesAddForm(request.POST,request.FILES)
        category_id  = request.POST.get('Category')
        if TopCategory.objects.filter(Category__id=category_id).exists():
            messages.error(request,'Already Exists')
            return redirect('siyyana_app:top_services')
        if form.is_valid():
            form.save()
            messages.success(request,'added successfully')
            return redirect('siyyana_app:top_services')
    else:
        form = TopServicesAddForm()
    context = {'form':form}
    return render(request,'add-top-services.html',context)

@login_required(login_url="siyyana_app:login")
def edit_topservices(request,id):
    update = TopCategory.objects.get(id=id)
    form = TopServicesAddForm(instance=update)
    if request.method == 'POST':
        form = TopServicesAddForm(request.POST,request.FILES,instance=update)
        if form.is_valid():
            form.save()
            messages.success(request,'Updated successfully')
            return redirect('siyyana_app:top_services')
    context = {'form':form}
    return render(request,'add-top-services.html',context)



@login_required(login_url="siyyana_app:login")
def delete_topservices(request,id):
    delete = TopCategory.objects.get(id=id)
    delete.delete()
    return redirect('siyyana_app:top_services')





@login_required(login_url="siyyana_app:login")
def employee_list(request):
    # Get form inputs
    country_id = request.GET.get('country')
    category_id = request.GET.get('category')
    state_id = request.GET.get('state')
    district_id = request.GET.get('district')
    subcategory_id = request.GET.get('subcategory')
    export_excel = request.GET.get('export_excel')

    # Start with all employees
    employees = CustomUser.objects.filter(user_type='Employee').order_by('-id')

    # Apply filters
    if country_id and country_id != 'All':
        employees = employees.filter(country__id=country_id)

    if category_id and category_id != 'All':
        employees = employees.filter(category__id=category_id)

    if state_id and state_id != 'All':
        employees = employees.filter(state__id=state_id)

    if district_id and district_id != 'All':
        employees = employees.filter(district__id=district_id)

    if subcategory_id and subcategory_id != 'All':
        employees = employees.filter(subcategory__id=subcategory_id)

    # Populate dropdown choices
    countries = Country.objects.all()
    categories = Category.objects.all()
    states = State.objects.all()
    districts = District.objects.all()
    subcategories = SubCategory.objects.all()


    if export_excel:
        return export_employees_to_excel(employees)

    context = {
        'users': employees,
        'countries': countries,
        'categories': categories,
        'states': states,
        'districts': districts,
        'subcategories': subcategories,
    }

    return render(request, 'employee.html', context)


def export_employees_to_excel(employees):
    # Create an in-memory output file for the Excel file
    output = BytesIO()

    # Create a workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Employees'

    # Define the column headers for employees
    headers = [
        'Name', 'Mobile Number', 'WhatsApp Number', 'About', 'Category', 'Subcategory', 
        'Charge', 'User Type', 'Date of Birth', 'Status',
        'Country', 'State', 'District', 'Preferred Work Location', 'ID Card Type', 'ID Card Number'
    ]

    # Write the headers to the first row
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Write the employee data to the Excel file
    for row_num, employee in enumerate(employees, 2):
        sheet.cell(row=row_num, column=1, value=employee.name)
        sheet.cell(row=row_num, column=2, value=employee.mobile_number)
        sheet.cell(row=row_num, column=3, value=employee.whatsapp_number)
        sheet.cell(row=row_num, column=4, value=employee.about)
        
        # Get Categories and Subcategories as comma-separated values
        categories = ", ".join([category.name for category in employee.category.all()])
        subcategories = ", ".join([subcategory.name for subcategory in employee.subcategory.all()])
        
        sheet.cell(row=row_num, column=5, value=categories)
        sheet.cell(row=row_num, column=6, value=subcategories)
        
        sheet.cell(row=row_num, column=7, value=employee.charge)
        sheet.cell(row=row_num, column=8, value=employee.user_type)
        sheet.cell(row=row_num, column=9, value=employee.date_of_birth)
        sheet.cell(row=row_num, column=10, value=employee.status)
        sheet.cell(row=row_num, column=11, value=employee.country.name if employee.country else 'No Country')
        sheet.cell(row=row_num, column=12, value=employee.state.name if employee.state else 'No State')
        sheet.cell(row=row_num, column=13, value=employee.district.name if employee.district else 'No District')
        sheet.cell(row=row_num, column=14, value=employee.prefered_work_location.name if employee.prefered_work_location else 'No Preferred Location')
        sheet.cell(row=row_num, column=15, value=employee.id_card_type)
        sheet.cell(row=row_num, column=16, value=employee.id_card_number)

    # Save the workbook to the BytesIO object
    workbook.save(output)
    output.seek(0)

    # Create an HTTP response with the Excel file
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'
    return response





def staff_status_change(request, id):
    statuss = CustomUser.objects.filter(id=id).first()
    if statuss.status == "Active":
        statuss.status = "Inactive"
    else:
        statuss.status = "Active"
    statuss.save()
    return redirect("siyyana_app:employee_list")

def view_profile(request,id):
    
    user = CustomUser.objects.get(id=id)
    # Initialize the form with the current user's data
    form = EmployeeViewForm(instance=user)

    # Render the form with disabled fields for view-only access
    context = {
        'form': form,
    }
    return render(request, 'view-profile.html', context)



@login_required(login_url="siyyana_app:login")
def employee_delete(request,id):
    data = CustomUser.objects.get(id=id)
    data.delete()
    messages.success(request,'Deleted successfully')
    return redirect('siyyana_app:employee_list')



@login_required(login_url="siyyana_app:login")
def user_list(request):
    # Get form inputs
    state_id = request.GET.get('state')
    district_id = request.GET.get('district')
    subcategory_id = request.GET.get('subcategory')
    export_excel = request.GET.get('export_excel')


    # Start with all users
    data = CustomUser.objects.filter(user_type='User').order_by('-id')

    # Apply filters


    if state_id and state_id != 'All':
        data = data.filter(state__id=state_id)

    if district_id and district_id != 'All':
        data = data.filter(district__id=district_id)

    if subcategory_id and subcategory_id != 'All':
        data = data.filter(subcategory__id=subcategory_id)

    # Populate dropdown choices
    countries = Country.objects.all()
    states = State.objects.all()
    districts = District.objects.all()

    if export_excel:
        return export_users_to_excel(data)


    context = {
        'users': data,
        'countries': countries,
        'states': states,
        'districts': districts,
    }

    return render(request, 'users.html', context)







def export_users_to_excel(employees):
    # Create an in-memory output file for the Excel file
    output = BytesIO()

    # Create a workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Users'

    # Define the column headers for employees
    headers = [
        'Name', 'Mobile Number', 'WhatsApp Number',
        'Country', 'State', 'District','Status'
    ]

    # Write the headers to the first row
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Write the employee data to the Excel file
    for row_num, employee in enumerate(employees, 2):
        sheet.cell(row=row_num, column=1, value=employee.name)
        sheet.cell(row=row_num, column=2, value=employee.mobile_number)
        sheet.cell(row=row_num, column=3, value=employee.whatsapp_number)
        sheet.cell(row=row_num, column=4, value=employee.country.name if employee.country else 'No Country')
        sheet.cell(row=row_num, column=5, value=employee.state.name if employee.state else 'No State')
        sheet.cell(row=row_num, column=6, value=employee.district.name if employee.district else 'No District')
        sheet.cell(row=row_num, column=7, value=employee.status)


    # Save the workbook to the BytesIO object
    workbook.save(output)
    output.seek(0)

    # Create an HTTP response with the Excel file
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=users.xlsx'
    return response





@login_required(login_url="siyyana_app:login")
def user_delete(request,id):
    data = CustomUser.objects.get(id=id)
    data.delete()
    messages.success(request,'Deleted successfully')
    return redirect('siyyana_app:user_list')




@login_required(login_url="siyyana_app:login")
def country(request):
    data = Country.objects.all().order_by('-id')
    context = {'data':data}
    return render(request,'country.html',context)




@login_required(login_url="siyyana_app:login")
def add_country(request):
    if request.method == 'POST':
        form = CountryAddForm(request.POST,request.FILES)
        name  = request.POST.get('name')
        if Country.objects.filter(name=name).exists():
            messages.error(request,'Already Exists')
            return redirect('siyyana_app:country')
        if form.is_valid():
            form.save()
            messages.success(request,'added successfully')
            return redirect('siyyana_app:country')
    else:
        form = CountryAddForm()
    context = {'form':form}
    return render(request,'add-country.html',context)





@login_required(login_url="siyyana_app:login")
def edit_country(request,id):
    update = Country.objects.get(id=id)
    if request.method == 'POST':
        form = CountryAddForm(request.POST,request.FILES,instance=update)
        if form.is_valid():
            form.save()
            messages.success(request,'Updated successfully')
            return redirect('siyyana_app:country')
    else:
        form = CountryAddForm(instance=update)
    context = {'form':form}
    return render(request,'add-country.html',context)



@login_required(login_url="siyyana_app:login")
def country_delete(request,id):
    data = Country.objects.get(id=id)
    data.delete()
    messages.success(request,'Deleted successfully')
    return redirect('siyyana_app:country')



# def import_excel_country(request):
#     if request.method == "POST":
        
#         if "excel_file" not in request.FILES:
#             print(request.POST)
#             messages.error(request, 'No file uploaded or incorrect file field name')
#             return redirect("siyyana_app:country")

#         excel_file = request.FILES["excel_file"]

#         try:
#             df = pd.read_excel(excel_file)

#             # Ensure that the DataFrame has the expected column
#             if "country" not in df.columns:
#                 messages.error(request, 'Excel file must contain a "country" column')
#                 return redirect("siyyana_app:country")

#             for index, row in df.iterrows():
#                 country_name = row["country"].strip()  # Ensure to strip any extra spaces
#                 Country.objects.update_or_create(
#                     name=country_name,
#                     defaults={"name": country_name},
#                 )
                
#             messages.success(request, "Successfully imported")
#         except Exception as e:
#             messages.error(request, f'Error processing the file: {e}')
        
#         return redirect("siyyana_app:country")









@login_required(login_url="siyyana_app:login")
def state(request):
    data = State.objects.all().order_by('-id')
    context = {'data':data}
    return render(request,'sate.html',context)

@login_required(login_url="siyyana_app:login")
def add_state(request):
    if request.method == 'POST':
        form = StateAddForm(request.POST,request.FILES)
        name  = request.POST.get('name')
        if State.objects.filter(name=name).exists():
            messages.error(request,'Already Exists')
            return redirect('siyyana_app:state')
        if form.is_valid():
            form.save()
            messages.success(request,'added successfully')
            return redirect('siyyana_app:state')
    else:
        form = StateAddForm()
    context = {'form':form}
    return render(request,'add-state.html',context)


@login_required(login_url="siyyana_app:login")
def edit_state(request,id):
    update = State.objects.get(id=id)
    if request.method == 'POST':
        form = StateAddForm(request.POST,request.FILES,instance=update)        
        if form.is_valid():
            form.save()
            messages.success(request,'updated successfully')
            return redirect('siyyana_app:state')
    else:
        form = StateAddForm(instance=update)
    context = {'form':form}
    return render(request,'add-state.html',context)



@login_required(login_url="siyyana_app:login")
def state_delete(request,id):
    data = State.objects.get(id=id)
    data.delete()
    messages.success(request,'Deleted successfully')
    return redirect('siyyana_app:state')



@login_required(login_url="siyyana_app:login")
def district(request):
    data = District.objects.all().order_by('-id')
    context = {'data':data}
    return render(request,'district.html',context)



@login_required(login_url="siyyana_app:login")
def add_district(request):
    if request.method == 'POST':
        form = DistrictAddForm(request.POST,request.FILES)
        name  = request.POST.get('name')
        if District.objects.filter(name=name).exists():
            messages.error(request,'Already Exists')
            return redirect('siyyana_app:district')
        if form.is_valid():
            form.save()
            messages.success(request,'added successfully')
            return redirect('siyyana_app:district')
    else:
        form = DistrictAddForm()
    context = {'form':form}
    return render(request,'add-district.html',context)



@login_required(login_url="siyyana_app:login")
def edit_district(request,id):
    update = District.objects.get(id=id)
    if request.method == 'POST':
        form = DistrictAddForm(request.POST,request.FILES,instance=update)
        if form.is_valid():
            form.save()
            messages.success(request,'Updated successfully')
            return redirect('siyyana_app:district')
    else:
        form = DistrictAddForm(instance=update)
    context = {'form':form}
    return render(request,'add-district.html',context)





@login_required(login_url="siyyana_app:login")
def district_delete(request,id):
    data = District.objects.get(id=id)
    data.delete()
    messages.success(request,'Deleted successfully')
    return redirect('siyyana_app:district')

